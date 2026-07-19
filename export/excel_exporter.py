"""
export/excel_exporter.py
========================
Writes the final attendance DataFrame to a beautifully styled
.xlsx file using openpyxl.

All visual concerns (colours, fonts, borders, column widths) live
here and nowhere else. The rest of the codebase is style-agnostic.
"""

import pandas as pd
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from config import EXCEL_SHEET_NAME


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

_COL_DATE_BG = "4F81BD"       # Blue header for the Date column
_COL_EMP_EVEN = "DCE6F1"      # Alternating employee header backgrounds
_COL_EMP_ODD = "E9EDF4"
_COL_ROW_EVEN = "F7F9FC"      # Alternating data row background

_MIN_COL_WIDTH = 10


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_to_excel(
    final_df: pd.DataFrame,
    unique_employees: list[str],
    output_path: str,
) -> None:
    """
    Write *final_df* to *output_path* as a formatted Excel workbook.

    Parameters
    ----------
    final_df:
        Multi-column pivot DataFrame produced by
        ``attendance.processor.build_attendance_dataframe``.
    unique_employees:
        Ordered list of employee label strings (same order as the
        columns in *final_df*).
    output_path:
        Destination ``.xlsx`` filename / path.
    """
    print("\nCreating beautiful attendance sheet…")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Write data starting at row 3 so rows 1–2 are free for custom headers.
        final_df.to_excel(
            writer,
            index=True,
            header=False,
            startrow=2,
            sheet_name=EXCEL_SHEET_NAME,
        )

        wb = writer.book
        ws = writer.sheets[EXCEL_SHEET_NAME]

        _apply_styles = _StyleKit()

        _write_date_header(ws, _apply_styles)
        _write_employee_headers(ws, unique_employees, _apply_styles)
        _style_data_rows(ws, _apply_styles)
        _auto_column_widths(ws)
        _set_row_heights(ws)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

class _StyleKit:
    """Reusable openpyxl style objects (created once, applied many times)."""

    font_header_top = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_header_sub = Font(name="Calibri", size=10, bold=True, color="333333")
    font_data = Font(name="Calibri", size=11)

    fill_date = PatternFill(start_color=_COL_DATE_BG, end_color=_COL_DATE_BG, fill_type="solid")
    fill_emp_even = PatternFill(start_color=_COL_EMP_EVEN, end_color=_COL_EMP_EVEN, fill_type="solid")
    fill_emp_odd = PatternFill(start_color=_COL_EMP_ODD, end_color=_COL_EMP_ODD, fill_type="solid")
    fill_row_even = PatternFill(start_color=_COL_ROW_EVEN, end_color=_COL_ROW_EVEN, fill_type="solid")

    _thin = Side(border_style="thin", color="D3D3D3")
    border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_date_header(ws, sk: _StyleKit) -> None:
    """Merge rows 1–2 in column A and write the 'Date' label."""
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    cell = ws.cell(row=1, column=1)
    cell.value = "Date"
    cell.fill = sk.fill_date
    cell.font = sk.font_header_top
    cell.alignment = sk.align_center
    cell.border = sk.border


def _write_employee_headers(
    ws,
    unique_employees: list[str],
    sk: _StyleKit,
) -> None:
    """Write merged employee name cells (row 1) and IN/Out/Hours sub-headers (row 2)."""
    current_col = 2

    for idx, emp in enumerate(unique_employees):
        fill = sk.fill_emp_even if idx % 2 == 0 else sk.fill_emp_odd

        # ── Employee name (spans 3 columns) ──────────────────────────────
        ws.merge_cells(
            start_row=1,
            start_column=current_col,
            end_row=1,
            end_column=current_col + 2,
        )
        emp_cell = ws.cell(row=1, column=current_col)
        emp_cell.value = emp
        emp_cell.font = Font(name="Calibri", size=11, bold=True, color="000000")
        emp_cell.fill = fill
        emp_cell.alignment = sk.align_center
        emp_cell.border = sk.border

        # ── Sub-headers: IN / Out / Hours ─────────────────────────────────
        sub_headers = [("IN", current_col), ("Out", current_col + 1), ("Hours", current_col + 2)]
        for label, col in sub_headers:
            cell = ws.cell(row=2, column=col)
            cell.value = label
            cell.font = sk.font_header_sub
            cell.fill = fill
            cell.alignment = sk.align_center
            cell.border = sk.border

        current_col += 3


def _style_data_rows(ws, sk: _StyleKit) -> None:
    """Apply font, alignment, border, and alternating fill to all data rows."""
    for row_idx in range(4, ws.max_row + 1):
        is_even = row_idx % 2 == 0
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = sk.font_data
            cell.alignment = sk.align_center
            cell.border = sk.border
            if is_even:
                cell.fill = sk.fill_row_even


def _auto_column_widths(ws) -> None:
    """Expand each column to fit its longest value (plus a small padding)."""
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = max(max_length + 4, _MIN_COL_WIDTH)


def _set_row_heights(ws) -> None:
    """Give the two header rows a comfortable fixed height."""
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
