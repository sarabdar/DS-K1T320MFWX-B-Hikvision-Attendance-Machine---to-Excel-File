import json
import requests
from time import sleep
from datetime import datetime
import calendar
from requests.auth import HTTPDigestAuth
import pandas as pd
import os

# openpyxl styling imports
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter



from dotenv import load_dotenv

load_dotenv()


# ==========================================
# DEVICE CONFIGURATION
# ==========================================

DEVICE_IP = os.getenv("DEVICE_IP")

USERNAME = os.getenv("USERNAME")
PASSWORD = os.getenv("PASSWORD")

url = (
    f"http://{DEVICE_IP}"
    f"/ISAPI/AccessControl/AcsEvent?format=json"
)

headers = {
    "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64)",
    "Content-Type": "application/json"
}

# ==========================================
# AUTOMATIC MONTH SELECTION
# ==========================================

current_year = datetime.now().year
current_month = datetime.now().month

current_month_name = calendar.month_name[current_month]

print(
    f"\nYou want to make Excel attendance sheet "
    f"for {current_month_name} {current_year}?"
)

choice = input(
    "Type y for yes and n for no: "
).strip().lower()

# ==========================================
# CURRENT MONTH
# ==========================================

if choice == "y":

    selected_month = current_month
    selected_year = current_year

# ==========================================
# CUSTOM MONTH
# ==========================================

else:

    print("\nOkay, which month?")
    print("Example: april")

    user_month = input(
        "Month name: "
    ).strip().lower()

    month_mapping = {
        "january": 1,
        "february": 2,
        "march": 3,
        "april": 4,
        "may": 5,
        "june": 6,
        "july": 7,
        "august": 8,
        "september": 9,
        "october": 10,
        "november": 11,
        "december": 12
    }

    if user_month not in month_mapping:

        print("\nInvalid month entered.")
        exit()

    selected_month = month_mapping[user_month]

    year_input = input(
        f"Enter year "
        f"(press Enter for {current_year}): "
    ).strip()

    selected_year = (
        int(year_input)
        if year_input
        else current_year
    )

# ==========================================
# BUILD DYNAMIC DATE RANGE
# ==========================================

month_name = calendar.month_name[selected_month]

last_day = calendar.monthrange(
    selected_year,
    selected_month
)[1]

START_TIME = (
    f"{selected_year}-{selected_month:02d}-01"
    f"T00:00:00+05:00"
)

END_TIME = (
    f"{selected_year}-{selected_month:02d}-{last_day:02d}"
    f"T23:59:59+05:00"
)

OUTPUT_EXCEL = (
    f"Hikvision_Attendance_"
    f"{month_name}_{selected_year}.xlsx"
)

print("\n==========================================")
print(f"Generating sheet for: {month_name} {selected_year}")
print(f"Start Time: {START_TIME}")
print(f"End Time:   {END_TIME}")
print(f"Output:     {OUTPUT_EXCEL}")
print("==========================================\n")

# ==========================================
# CREATE SESSION
# ==========================================

session = requests.Session()

session.proxies = {
    "http": None,
    "https": None
}

# ==========================================
# FETCH EVENTS
# ==========================================

raw_events = []

position = 0
page_size = 500

print(
    f"Starting data extraction from "
    f"Hikvision device at {DEVICE_IP}..."
)

while True:

    payload = {
        "AcsEventCond": {
            "searchID": "1",
            "searchResultPosition": position,
            "maxResults": page_size,
            "startTime": START_TIME,
            "endTime": END_TIME,
            "major": 0,
            "minor": 0
        }
    }

    try:

        response = session.post(
            url,
            auth=HTTPDigestAuth(
                USERNAME,
                PASSWORD
            ),
            headers=headers,
            data=json.dumps(payload),
            timeout=15
        )

        response.raise_for_status()

        data = response.json()

        acs_event = data.get("AcsEvent", {})

        info_list = acs_event.get(
            "InfoList",
            []
        )

        status_strg = acs_event.get(
            "responseStatusStrg",
            "OK"
        )

        total_matches = acs_event.get(
            "totalMatches",
            0
        )

        if not info_list:
            break

        print(
            f"Fetched records {position} to "
            f"{position + len(info_list)} "
            f"(Total Available: {total_matches})"
        )

        # ==========================================
        # KEEP ONLY REAL ATTENDANCE EVENTS
        # ==========================================

        for log in info_list:

            if (
                "employeeNoString" in log
                and log.get("time")
                and log.get("attendanceStatus")
            ):

                raw_events.append(log)

        if status_strg != "MORE":
            break

        position += len(info_list)

        sleep(0.1)

    except Exception as e:

        print(
            f"\nError occurred during "
            f"pagination loop: {e}"
        )

        break

# ==========================================
# PROCESS ATTENDANCE
# ==========================================

if raw_events:

    print(
        f"\nProcessing "
        f"{len(raw_events)} attendance logs..."
    )

    df_raw = pd.DataFrame(raw_events)


    # ==========================================
    # CONVERT DATETIME
    # ==========================================

    df_raw['datetime'] = pd.to_datetime(
        df_raw['time']
    )

    # Extract date
    df_raw['Date'] = (
        df_raw['datetime']
        .dt.strftime('%Y-%m-%d')
    )

    # Employee column
    df_raw['Employee_Col'] = (
        df_raw['name'].fillna('Unknown')
        + ' '
        + df_raw['employeeNoString']
    )

    # ==========================================
    # FORMAT TIME
    # ==========================================

    df_raw['FormattedTime'] = (
        df_raw['datetime']
        .dt.strftime('%I:%M%p')
        .str.lstrip('0')
        .str.lower()
    )

    # ==========================================
    # CREATE ATTENDANCE TABLE
    # ==========================================

    attendance_rows = []

    grouped_data = df_raw.groupby(
        ['Date', 'Employee_Col']
    )

    for (date_val, employee), group in grouped_data:

        check_in_times = []
        check_out_times = []

        for _, row in group.iterrows():

            status = str(
                row.get('attendanceStatus', '')
            ).lower()

            if status == 'checkin':

                check_in_times.append(
                    row['datetime']
                )

            elif status == 'checkout':

                check_out_times.append(
                    row['datetime']
                )

        # Earliest check-in
        if check_in_times:

            in_time = min(check_in_times)

            in_time_str = (
                in_time.strftime('%I:%M%p')
                .lstrip('0')
                .lower()
            )

        else:

            in_time_str = ''

        # Latest check-out
        if check_out_times:

            out_time = max(check_out_times)

            out_time_str = (
                out_time.strftime('%I:%M%p')
                .lstrip('0')
                .lower()
            )

        else:

            out_time_str = ''

        # ==========================================
        # CALCULATE SHIFT HOURS
        # ==========================================

        hours_str = ''

        if check_in_times and check_out_times:

            # If checkout is next day
            if out_time < in_time:

                out_time = out_time + pd.Timedelta(days=1)

            shift_duration = out_time - in_time

            total_seconds = int(
                shift_duration.total_seconds()
            )

            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60

            hours_str = f"{hours}h {minutes}m"
        total_minutes = 0

        if check_in_times and check_out_times:

            if out_time < in_time:

                out_time = out_time + pd.Timedelta(days=1)

            shift_duration = out_time - in_time

            total_minutes = int(
                shift_duration.total_seconds() / 60
            )

            hours = total_minutes // 60
            minutes = total_minutes % 60

            hours_str = f"{hours}h {minutes}m"

        else:
            hours_str = ''
        attendance_rows.append({
    'Date': date_val,
    'Employee_Col': employee,
    'IN': in_time_str,
    'Out': out_time_str,
    'Hours': hours_str,
    'TotalMinutes': total_minutes
})

    grouped = pd.DataFrame(attendance_rows)

    # ==========================================
    # CREATE FINAL TABLE
    # ==========================================

    unique_dates = sorted(
        grouped['Date'].unique()
    )

    unique_employees = sorted(
        grouped['Employee_Col'].unique()
    )

    columns = []

    for emp in unique_employees:

        columns.append((emp, 'IN'))
        columns.append((emp, 'Out'))
        columns.append((emp, 'Hours'))

    multi_cols = pd.MultiIndex.from_tuples(
        columns
    )

    final_df = pd.DataFrame(
        index=unique_dates,
        columns=multi_cols
    )

    for _, row in grouped.iterrows():

        date_val = row['Date']
        emp = row['Employee_Col']

        final_df.loc[
            date_val,
            (emp, 'IN')
        ] = row['IN']

        final_df.loc[
            date_val,
            (emp, 'Out')
        ] = row['Out']

        final_df.loc[
            date_val,
            (emp, 'Hours')] = row['Hours']

    final_df = final_df.fillna('')

# ==========================================
# ADD TOTAL HOURS ROW
# ==========================================

    totals_row = {}

    for emp in unique_employees:

        emp_data = grouped[
            grouped['Employee_Col'] == emp
        ]

        total_minutes = emp_data[
            'TotalMinutes'
        ].sum()

        total_hours = total_minutes // 60
        remaining_minutes = total_minutes % 60

        total_hours_str = (
            f"{total_hours}h "
            f"{remaining_minutes}m"
        )

        totals_row[(emp, 'IN')] = ''
        totals_row[(emp, 'Out')] = 'TOTAL'
        totals_row[(emp, 'Hours')] = total_hours_str

    # Add totals row at bottom
    final_df.loc['TOTAL'] = totals_row

    # ==========================================
    # EXPORT EXCEL
    # ==========================================

    print(
        "\nCreating beautiful attendance sheet..."
    )

    with pd.ExcelWriter(
        OUTPUT_EXCEL,
        engine='openpyxl'
    ) as writer:

        # Write data without headers
        final_df.to_excel(
            writer,
            index=True,
            header=False,
            startrow=2,
            sheet_name="Attendance Sheet"
        )

        workbook = writer.book

        worksheet = writer.sheets[
            "Attendance Sheet"
        ]

        # ==========================================
        # STYLES
        # ==========================================

        font_header_top = Font(
            name="Calibri",
            size=11,
            bold=True,
            color="FFFFFF"
        )

        font_header_sub = Font(
            name="Calibri",
            size=10,
            bold=True,
            color="333333"
        )

        font_data = Font(
            name="Calibri",
            size=11
        )

        fill_date = PatternFill(
            start_color="4F81BD",
            end_color="4F81BD",
            fill_type="solid"
        )

        fill_emp_even = PatternFill(
            start_color="DCE6F1",
            end_color="DCE6F1",
            fill_type="solid"
        )

        fill_emp_odd = PatternFill(
            start_color="E9EDF4",
            end_color="E9EDF4",
            fill_type="solid"
        )

        fill_data_even = PatternFill(
            start_color="F7F9FC",
            end_color="F7F9FC",
            fill_type="solid"
        )

        thin_side = Side(
            border_style="thin",
            color="D3D3D3"
        )

        border_all = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )

        align_center = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # ==========================================
        # DATE HEADER
        # ==========================================

        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=1
        )

        date_cell = worksheet.cell(
            row=1,
            column=1
        )

        date_cell.value = "Date"
        date_cell.fill = fill_date
        date_cell.font = font_header_top
        date_cell.alignment = align_center
        date_cell.border = border_all

        # ==========================================
        # EMPLOYEE HEADERS
        # ==========================================

        current_col = 2

        for idx, emp in enumerate(
            unique_employees
        ):

            fill = (
                fill_emp_even
                if idx % 2 == 0
                else fill_emp_odd
            )

            worksheet.merge_cells(
                start_row=1,
                start_column=current_col,
                end_row=1,
                end_column=current_col + 2
            )

            emp_cell = worksheet.cell(
                row=1,
                column=current_col
            )

            emp_cell.value = emp

            emp_cell.font = Font(
                name="Calibri",
                size=11,
                bold=True,
                color="000000"
            )

            emp_cell.fill = fill
            emp_cell.alignment = align_center
            emp_cell.border = border_all

            # IN
            in_cell = worksheet.cell(
                row=2,
                column=current_col
            )

            in_cell.value = "IN"
            in_cell.font = font_header_sub
            in_cell.fill = fill
            in_cell.alignment = align_center
            in_cell.border = border_all

            # OUT
            out_cell = worksheet.cell(
                row=2,
                column=current_col + 1
            )

            out_cell.value = "Out"
            out_cell.font = font_header_sub
            out_cell.fill = fill
            out_cell.alignment = align_center
            out_cell.border = border_all

            # HOURS
            hours_cell = worksheet.cell(
                row=2,
                column=current_col + 2
            )

            hours_cell.value = "Hours"
            hours_cell.font = font_header_sub
            hours_cell.fill = fill
            hours_cell.alignment = align_center
            hours_cell.border = border_all

            current_col += 3

        # ==========================================
        # STYLE DATA ROWS
        # ==========================================

        for row_idx in range(
            4,
            worksheet.max_row + 1
        ):

            is_even_row = (
                row_idx % 2 == 0
            )

            for col_idx in range(
                1,
                worksheet.max_column + 1
            ):

                cell = worksheet.cell(
                    row=row_idx,
                    column=col_idx
                )

                cell.font = font_data
                cell.alignment = align_center
                cell.border = border_all

                if is_even_row:
                    cell.fill = fill_data_even

        # ==========================================
        # AUTO COLUMN WIDTH
        # ==========================================

        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = (
                get_column_letter(
                    column_cells[0].column
                )
            )

            for cell in column_cells:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except:
                    pass

            adjusted_width = max(
                max_length + 4,
                10
            )

            worksheet.column_dimensions[
                column_letter
            ].width = adjusted_width

        # ==========================================
        # ROW HEIGHTS
        # ==========================================

        worksheet.row_dimensions[1].height = 28
        worksheet.row_dimensions[2].height = 22

    print("\n==========================================")
    print("DONE SUCCESSFULLY!")
    print(f"Excel file created: {OUTPUT_EXCEL}")
    print("==========================================")

else:

    print(
        "\nNo attendance logs found "
        "for the selected month."
    )